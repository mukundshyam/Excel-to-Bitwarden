import json
import urllib.parse, urllib.error, urllib.request
import openpyxl
import tldextract


Instructions = """Before starting, make sure to follow these steps. 
1. Make sure your file is in the .xslx format. If it's not, make sure to export/save as a .xslx file. 
2. Keep an eye on where this app is installed. That will be where your output sheet will appear. 
3. Make sure your file is formatted correctly. 
The first column (Column A) should be left empty. 
The second column (B) should contain all your usernames. 
The third column (C) should contain all your passwords. 
The fourth column (D) should contain the labels for your login information (for example, 'google' or 'hdfc').
Make sure these cells contain the following labels:
    Cell A1 = url
    Cell B1 = username
    Cell C1 = password
    Cell D1 = name
4. If your sheet has more than 100 logins, please split it into multiple sheets, and combine it later. I'm limited to 100 API calls/day. :) """

print(Instructions)
path = input("Enter filepath - ")

try:
    wbobj = openpyxl.load_workbook(path)
except:
    print("Sorry, that didn't work. Try to remove quotation marks, and paste the correct path for your .xlsx file.")
    quit()

sheetobj = wbobj.active
row = 2
column = 4

apiKey = "AIzaSyBCLq_yTsVh0s-gS1A9nYAAPQk17LPt0sE"
customsearch = "30d276c4177f84d8f"

val = sheetobj.cell(row=row, column=column).value

while val != None:
    query = val.replace(" ", "%20")
    try:
        jsonobj = urllib.request.urlopen("https://www.googleapis.com/customsearch/v1?key=" + apiKey + "&cx=" + customsearch + "&q=" + query + "%20website&num=1")
    except:
        print("Out of API calls. Please try again tomorrow.")
        quit()
    obj = jsonobj.read()
    parses = json.loads(obj)
    dic = parses["items"]
    out = dic[0]["link"]
    links = tldextract.extract(out)
    formatted = '{}.{}'.format(links.domain, links.suffix)
    print(val, formatted)
    sheetobj.cell(row=row, column=1).value = formatted
    row = row+1
    val = sheetobj.cell(row=row, column=column).value

wbobj.save("output.xlsx")

NextSteps = """Next up, make sure to take the output book (saved as output.xlsx), and open it in Excel. 
Then, click on Save As, and save it as a .csv file.
Next, make a Bitwarden account, and import this .csv file into your web vault. (use this link to help: https://bitwarden.com/help/import-from-firefox/)
When you import, make sure to set the file format as 'firefox (.csv).'
Voila! You now have a fully fledged Bitwarden vault!"""

print(NextSteps)